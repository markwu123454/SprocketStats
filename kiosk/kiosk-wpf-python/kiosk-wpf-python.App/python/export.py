"""
export_raw.py – Emergency Excel export of aggregated team data.

Pulls directly from TBA, Statbotics, and downloaded scouting data so
the spreadsheet is usable even when the main calculator pipeline is broken.
"""

import json
import statistics
from typing import Optional

import requests
import statbotics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from context import ctx
from calculator import (
    DownloadedData,
    Match,
    MatchAlliance,
    MatchScoutingEntry,
    AllianceScoreBreakdown2026,
    StatboticsMatch,
    StatboticsAllianceData,
    StatboticsPred,
    ScoreAction,
    ClimbAction,
    ShootingAction,
    parse_team_key,
    compute_time_percentages,
    _STATE_LABELS,
)

_sb = statbotics.Statbotics()

# ── Style constants ──────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
_SUBHEADER_FONT = Font(name="Arial", bold=True, size=10)
_DATA_FONT = Font(name="Arial", size=10)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
)
_PCT_FMT = "0.0%"
_DEC2_FMT = "0.00"
_DEC1_FMT = "0.0"
_INT_FMT = "0"


def _safe_mean(vals):
    return round(statistics.mean(vals), 2) if vals else None


def _safe_stdev(vals):
    return round(statistics.stdev(vals), 2) if vals and len(vals) > 1 else None


def _safe_rate(ones_zeros):
    return round(sum(ones_zeros) / len(ones_zeros), 3) if ones_zeros else None


def _style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, max_col, min_width=10, max_width=22):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    best = max(best, min(len(str(cell.value)) + 3, max_width))
        ws.column_dimensions[letter].width = best


# ── Core aggregation (no dependency on run_calculation) ──────────────────

def _aggregate_teams(event_key: str, downloaded_data: DownloadedData, tba_data: list[Match], sb_data: list[StatboticsMatch]):
    """Build a dict[team_num -> row_dict] from raw sources."""

    teams: dict[int, dict] = {}

    # ── 1. Discover all teams from TBA matches ──
    for m in tba_data:
        for color in ("red", "blue"):
            al: MatchAlliance = getattr(m.alliances, color)
            for tk in al.team_keys:
                tn = parse_team_key(tk)
                teams.setdefault(tn, {})

    # ── 2. TBA qual breakdowns ──
    quals = [m for m in tba_data if m.comp_level == "qm" and m.score_breakdown is not None]

    tba_rp: dict[int, int] = {}
    tba_played: dict[int, int] = {}
    tba_auto_pts: dict[int, list] = {}
    tba_teleop_pts: dict[int, list] = {}
    tba_total_pts: dict[int, list] = {}
    tba_tower_pts: dict[int, list] = {}
    tba_foul_pts: dict[int, list] = {}
    tba_hub_pts: dict[int, list] = {}
    tba_energized: dict[int, int] = {}
    tba_supercharged: dict[int, int] = {}
    tba_traversal: dict[int, int] = {}

    for m in quals:
        for color in ("red", "blue"):
            al: MatchAlliance = getattr(m.alliances, color)
            bd: AllianceScoreBreakdown2026 = getattr(m.score_breakdown, color)
            for tk in al.team_keys:
                tn = parse_team_key(tk)
                tba_played[tn] = tba_played.get(tn, 0) + 1
                tba_rp[tn] = tba_rp.get(tn, 0) + bd.rp
                tba_auto_pts.setdefault(tn, []).append(bd.totalAutoPoints / 3)
                tba_teleop_pts.setdefault(tn, []).append(bd.totalTeleopPoints / 3)
                tba_total_pts.setdefault(tn, []).append(bd.totalPoints / 3)
                tba_tower_pts.setdefault(tn, []).append(bd.totalTowerPoints / 3)
                tba_foul_pts.setdefault(tn, []).append(bd.foulPoints / 3)
                tba_hub_pts.setdefault(tn, []).append(bd.hubScore.totalPoints / 3)
                if bd.energizedAchieved:
                    tba_energized[tn] = tba_energized.get(tn, 0) + 1
                if bd.superchargedAchieved:
                    tba_supercharged[tn] = tba_supercharged.get(tn, 0) + 1
                if bd.traversalAchieved:
                    tba_traversal[tn] = tba_traversal.get(tn, 0) + 1

    # ── 3. Statbotics predicted RP ──
    sb_rp_pred: dict[int, float] = {}
    sb_pred_score: dict[int, list] = {}
    sb_quals_m = [m for m in sb_data if m.comp_level == "qm" and m.pred is not None]
    for m in sb_quals_m:
        for color in ("red", "blue"):
            ad: StatboticsAllianceData = getattr(m.alliances, color)
            p: StatboticsPred = m.pred
            win_prob = p.red_win_prob if color == "red" else (1 - p.red_win_prob)
            rp1 = p.red_rp_1 if color == "red" else p.blue_rp_1
            rp2 = p.red_rp_2 if color == "red" else p.blue_rp_2
            rp3 = (p.red_rp_3 if color == "red" else p.blue_rp_3) or 0.0
            pred_score = p.red_score if color == "red" else p.blue_score
            expected_rp = 3.0 * win_prob + rp1 + rp2 + rp3
            for tk in ad.team_keys:
                tn = parse_team_key(tk)
                sb_rp_pred[tn] = sb_rp_pred.get(tn, 0.0) + expected_rp
                sb_pred_score.setdefault(tn, []).append(pred_score / 3)

    # ── 4. Scouting data ──
    scouting = [e for e in downloaded_data.match_scouting if e.event_key == event_key]

    sc_total_fuel: dict[int, list] = {}
    sc_auto_fuel: dict[int, list] = {}
    sc_teleop_fuel: dict[int, list] = {}
    sc_accuracy: dict[int, list] = {}
    sc_auto_climb_ok: dict[int, list] = {}
    sc_endgame_climb_ok: dict[int, list] = {}
    sc_climb_pts: dict[int, list] = {}
    sc_endgame_levels: dict[int, list] = {}
    sc_skill: dict[int, list] = {}
    sc_defense: dict[int, list] = {}
    sc_speed: dict[int, list] = {}
    sc_faults: dict[int, list] = {}
    sc_roles: dict[int, list] = {}
    sc_time_pcts: dict[int, dict[str, list]] = {}
    sc_matches: dict[int, int] = {}

    TOWER_PTS = {"L1": 10, "L2": 20, "L3": 30}
    subphase_teleop = {"transition", "shift_1", "shift_2", "shift_3", "shift_4", "endgame"}

    for entry in scouting:
        tn = int(entry.team)
        actions = entry.data.actions
        pm = entry.data.postmatch

        sc_matches[tn] = sc_matches.get(tn, 0) + 1

        total_scored = 0
        auto_scored = 0
        teleop_scored = 0
        total_shots = 0
        auto_climb = 0.0
        endgame_climb = 0.0
        climb_pts = 0.0
        endgame_level = 0

        for a in actions:
            if isinstance(a, ShootingAction):
                total_shots += 1
            elif isinstance(a, ScoreAction):
                total_scored += a.score
                if a.subPhase == "auto":
                    auto_scored += a.score
                elif a.subPhase in subphase_teleop:
                    teleop_scored += a.score
            elif isinstance(a, ClimbAction):
                if a.phase == "auto":
                    if a.success:
                        auto_climb = 1.0
                        climb_pts += 15.0
                else:
                    if a.success:
                        endgame_climb = 1.0
                        lvl_pts = TOWER_PTS.get(a.level, 10)
                        climb_pts += lvl_pts
                        endgame_level = {"L1": 1, "L2": 2, "L3": 3}.get(a.level, 0)

        sc_total_fuel.setdefault(tn, []).append(float(total_scored))
        sc_auto_fuel.setdefault(tn, []).append(float(auto_scored))
        sc_teleop_fuel.setdefault(tn, []).append(float(teleop_scored))
        sc_accuracy.setdefault(tn, []).append(total_scored / total_shots if total_shots > 0 else 0.0)
        sc_auto_climb_ok.setdefault(tn, []).append(auto_climb)
        sc_endgame_climb_ok.setdefault(tn, []).append(endgame_climb)
        sc_climb_pts.setdefault(tn, []).append(climb_pts)
        if endgame_level > 0:
            sc_endgame_levels.setdefault(tn, []).append(endgame_level)

        sc_skill.setdefault(tn, []).append(pm.skill)
        sc_defense.setdefault(tn, []).append(pm.defenseSkill)
        sc_speed.setdefault(tn, []).append(pm.speed)
        sc_roles.setdefault(tn, []).append(pm.role)

        fault_fields = [
            "jam", "other", "brownout", "disabled", "failed_auto",
            "immobilized", "disconnected", "erratic_driving", "structural_failure",
        ]
        fault_count = sum(1 for f in fault_fields if getattr(pm.faults, f))
        sc_faults.setdefault(tn, []).append(fault_count)

        time_pcts = compute_time_percentages(actions)
        sc_time_pcts.setdefault(tn, {s: [] for s in _STATE_LABELS})
        for s in _STATE_LABELS:
            sc_time_pcts[tn][s].append(time_pcts.get(s, 0.0))

    # ── 5. Assemble rows ──
    for tn in sorted(teams.keys()):
        mp = tba_played.get(tn, 0)
        sm = sc_matches.get(tn, 0)
        roles = sc_roles.get(tn, [])
        primary_role = max(set(roles), key=roles.count) if roles else None

        levels = sc_endgame_levels.get(tn, [])
        best_level = max(set(levels), key=levels.count) if levels else None

        row = {
            # TBA
            "tba_matches_played": mp,
            "tba_rp_total": tba_rp.get(tn, 0),
            "tba_rp_avg": round(tba_rp.get(tn, 0) / mp, 2) if mp > 0 else None,
            "tba_auto_pts_avg": _safe_mean(tba_auto_pts.get(tn, [])),
            "tba_teleop_pts_avg": _safe_mean(tba_teleop_pts.get(tn, [])),
            "tba_total_pts_avg": _safe_mean(tba_total_pts.get(tn, [])),
            "tba_tower_pts_avg": _safe_mean(tba_tower_pts.get(tn, [])),
            "tba_hub_pts_avg": _safe_mean(tba_hub_pts.get(tn, [])),
            "tba_foul_pts_avg": _safe_mean(tba_foul_pts.get(tn, [])),
            "tba_energized_rate": round(tba_energized.get(tn, 0) / mp, 3) if mp > 0 else None,
            "tba_supercharged_rate": round(tba_supercharged.get(tn, 0) / mp, 3) if mp > 0 else None,
            "tba_traversal_rate": round(tba_traversal.get(tn, 0) / mp, 3) if mp > 0 else None,
            # Statbotics
            "sb_pred_rp_total": round(sb_rp_pred.get(tn, 0), 2) if tn in sb_rp_pred else None,
            "sb_pred_rp_avg": round(sb_rp_pred[tn] / max(mp, 1), 2) if tn in sb_rp_pred else None,
            "sb_pred_score_avg": _safe_mean(sb_pred_score.get(tn, [])),
            # Scouting
            "sc_matches_scouted": sm,
            "sc_total_fuel_avg": _safe_mean(sc_total_fuel.get(tn, [])),
            "sc_total_fuel_stdev": _safe_stdev(sc_total_fuel.get(tn, [])),
            "sc_auto_fuel_avg": _safe_mean(sc_auto_fuel.get(tn, [])),
            "sc_teleop_fuel_avg": _safe_mean(sc_teleop_fuel.get(tn, [])),
            "sc_accuracy_avg": _safe_mean(sc_accuracy.get(tn, [])),
            "sc_climb_pts_avg": _safe_mean(sc_climb_pts.get(tn, [])),
            "sc_auto_climb_rate": _safe_rate(sc_auto_climb_ok.get(tn, [])),
            "sc_endgame_climb_rate": _safe_rate(sc_endgame_climb_ok.get(tn, [])),
            "sc_best_climb_level": best_level,
            "sc_est_total_avg": (
                round((_safe_mean(sc_total_fuel.get(tn, [])) or 0) + (_safe_mean(sc_climb_pts.get(tn, [])) or 0), 2)
                if sm > 0 else None
            ),
            "sc_skill_avg": _safe_mean(sc_skill.get(tn, [])),
            "sc_defense_skill_avg": _safe_mean(sc_defense.get(tn, [])),
            "sc_speed_avg": _safe_mean(sc_speed.get(tn, [])),
            "sc_fault_rate": _safe_mean(sc_faults.get(tn, [])),
            "sc_primary_role": primary_role,
        }

        # Time percentages
        for s in _STATE_LABELS:
            row[f"sc_time_{s}_avg"] = _safe_mean(sc_time_pcts.get(tn, {}).get(s, []))

        teams[tn] = row

    return teams


# ── Column definitions for the spreadsheet ───────────────────────────────

COLUMNS = [
    ("Team", "team", _INT_FMT),
    # TBA
    ("TBA Matches", "tba_matches_played", _INT_FMT),
    ("TBA RP Total", "tba_rp_total", _INT_FMT),
    ("TBA RP Avg", "tba_rp_avg", _DEC2_FMT),
    ("TBA Auto Pts Avg*", "tba_auto_pts_avg", _DEC2_FMT),
    ("TBA Teleop Pts Avg*", "tba_teleop_pts_avg", _DEC2_FMT),
    ("TBA Total Pts Avg*", "tba_total_pts_avg", _DEC2_FMT),
    ("TBA Tower Pts Avg*", "tba_tower_pts_avg", _DEC2_FMT),
    ("TBA Hub Pts Avg*", "tba_hub_pts_avg", _DEC2_FMT),
    ("TBA Foul Pts Avg*", "tba_foul_pts_avg", _DEC2_FMT),
    ("TBA Energized Rate", "tba_energized_rate", _DEC2_FMT),
    ("TBA Supercharged Rate", "tba_supercharged_rate", _DEC2_FMT),
    ("TBA Traversal Rate", "tba_traversal_rate", _DEC2_FMT),
    # Statbotics
    ("SB Pred RP Total", "sb_pred_rp_total", _DEC2_FMT),
    ("SB Pred RP Avg", "sb_pred_rp_avg", _DEC2_FMT),
    ("SB Pred Score Avg*", "sb_pred_score_avg", _DEC2_FMT),
    # Scouting
    ("Scouted Matches", "sc_matches_scouted", _INT_FMT),
    ("SC Total Fuel Avg", "sc_total_fuel_avg", _DEC2_FMT),
    ("SC Total Fuel StDev", "sc_total_fuel_stdev", _DEC2_FMT),
    ("SC Auto Fuel Avg", "sc_auto_fuel_avg", _DEC2_FMT),
    ("SC Teleop Fuel Avg", "sc_teleop_fuel_avg", _DEC2_FMT),
    ("SC Accuracy Avg", "sc_accuracy_avg", _DEC2_FMT),
    ("SC Climb Pts Avg", "sc_climb_pts_avg", _DEC2_FMT),
    ("SC Auto Climb Rate", "sc_auto_climb_rate", _DEC2_FMT),
    ("SC Endgame Climb Rate", "sc_endgame_climb_rate", _DEC2_FMT),
    ("SC Best Climb Level", "sc_best_climb_level", _INT_FMT),
    ("SC Est Total Avg", "sc_est_total_avg", _DEC2_FMT),
    ("SC Skill Avg", "sc_skill_avg", _DEC2_FMT),
    ("SC Defense Skill Avg", "sc_defense_skill_avg", _DEC2_FMT),
    ("SC Speed Avg", "sc_speed_avg", _DEC2_FMT),
    ("SC Fault Rate", "sc_fault_rate", _DEC2_FMT),
    ("SC Primary Role", "sc_primary_role", None),
]

# Add time percentage columns
for _s in _STATE_LABELS:
    COLUMNS.append((f"SC Time {_s.title()} %", f"sc_time_{_s}_avg", _DEC1_FMT))


# ── Public entry point ───────────────────────────────────────────────────

def export_raw_excel(event_key: str, output_path: str = "raw_export.xlsx") -> str:
    """
    Fetch data from TBA, Statbotics, and scouting, aggregate per-team
    stats, and write to an Excel file. Returns the output path.
    """

    downloaded_data = DownloadedData(**ctx.downloaded_data)

    # Fetch TBA
    tba_resp = requests.get(
        f"https://www.thebluealliance.com/api/v3/event/{event_key}/matches",
        headers={"X-TBA-Auth-Key": ctx.TBA_API_KEY},
    )
    tba_data = [Match(**m) for m in tba_resp.json()] if tba_resp.status_code == 200 else []

    # Fetch Statbotics
    try:
        sb_data = [StatboticsMatch(**m) for m in _sb.get_matches(event=event_key)]
    except Exception:
        sb_data = []

    teams = _aggregate_teams(event_key, downloaded_data, tba_data, sb_data)

    # ── Build workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Summary"

    # Freeze top row + team column
    ws.freeze_panes = "B2"

    # Header row
    for ci, (label, _, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=ci, value=label)
    _style_header_row(ws, 1, len(COLUMNS))

    # Data rows (sorted by team number)
    for ri, tn in enumerate(sorted(teams.keys()), start=2):
        row = teams[tn]
        ws.cell(row=ri, column=1, value=tn)
        for ci, (_, key, fmt) in enumerate(COLUMNS[1:], start=2):
            val = row.get(key)
            cell = ws.cell(row=ri, column=ci, value=val)
            if fmt and val is not None:
                cell.number_format = fmt

    last_row = 1 + len(teams)
    _style_data_rows(ws, 2, last_row, len(COLUMNS))
    _auto_width(ws, len(COLUMNS))

    # Add a note about alliance-level TBA stats
    note_ws = wb.create_sheet("Notes")
    note_ws["A1"] = "Column Notes"
    note_ws["A1"].font = Font(name="Arial", bold=True, size=12)
    note_ws["A3"] = "* TBA point columns marked with * are alliance totals divided by 3."
    note_ws["A4"] = "  They approximate per-robot contribution but are not exact."
    note_ws["A6"] = "SC = Scouting data,  SB = Statbotics,  TBA = The Blue Alliance"
    note_ws["A8"] = "Time % columns show average % of match time spent in each state."
    note_ws["A10"] = f"Event: {event_key}"
    for r in range(1, 11):
        note_ws.cell(row=r, column=1).font = _DATA_FONT

    import os
    wb.save(output_path)
    return os.path.abspath(output_path)